"""
Motor de procesamiento de archivos Excel courier.
- Detección flexible de columnas (aliases)
- Normalización de nombres de ciudades (variantes → canónico)
- Detección y marcado amarillo de ciudades ambiguas
- Búsqueda case-insensitive en nombre de destinatario
- Matching exacto de patrones con regex
- Construcción de workbook con formato profesional
"""
import io
import re
import logging

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import BOGOTA_ALIASES, COLUMN_ALIASES, REQUIRED_COLUMNS
from city_normalize import normalize_city, is_ambiguous, normalize_city_series

logger = logging.getLogger(__name__)

# ── Estilos ───────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="D9E1F2")
NORMAL_FILL  = PatternFill("solid", start_color="FFFFFF")
AMBIG_FILL   = PatternFill("solid", start_color="FFD966")   # ← amarillo ciudad ambigua
AMBIG_FONT   = Font(name="Arial", size=10, bold=True, color="7F6000")
CELL_FONT    = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="BDD7EE")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COL_WIDTHS = {
    "Guia#": 22, "Nombre Destinatario": 30, "Dirección": 35,
    "Tel-1": 16, "Detalle": 40, "Ciudad": 20, "Piezas": 8, "Peso-2": 10,
}


# ── Detección de columnas ─────────────────────────────────────────────────────

def detect_columns(df: pd.DataFrame) -> dict[str, str]:
    """
    Mapea nombres canónicos → nombres reales en el DataFrame.
    Lanza ValueError si falta alguna columna requerida.
    """
    actual_lower = {c.lower().strip(): c for c in df.columns}
    mapping: dict[str, str] = {}

    for canonical, aliases in COLUMN_ALIASES.items():
        found = actual_lower.get(canonical.lower())
        if not found:
            for alias in aliases:
                found = actual_lower.get(alias.lower())
                if found:
                    break
        if found:
            mapping[canonical] = found

    missing = [c for c in REQUIRED_COLUMNS if c not in mapping]
    if missing:
        available = ", ".join(df.columns.tolist())
        raise ValueError(
            f"No se encontraron las columnas requeridas: *{', '.join(missing)}*\n\n"
            f"Columnas disponibles en el archivo:\n`{available}`\n\n"
            f"Renombra las columnas o contacta al administrador para agregar el alias."
        )

    logger.info(f"Columnas detectadas: {mapping}")
    return mapping


def normalize_df(df: pd.DataFrame, col_map: dict[str, str]) -> pd.DataFrame:
    """
    Renombra columnas al nombre canónico, limpia datos
    y normaliza los nombres de ciudad.
    """
    rename = {v: k for k, v in col_map.items()}
    df = df.rename(columns=rename)
    known = [c for c in COLUMN_ALIASES.keys() if c in df.columns]
    df = df[known].copy()
    df.fillna("", inplace=True)
    df = df.astype(str)
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    # Normalizar ciudades → nombre canónico unificado
    if "Ciudad" in df.columns:
        df["Ciudad"] = normalize_city_series(df["Ciudad"])

    return df


# ── Clasificación ─────────────────────────────────────────────────────────────

def is_bogota(ciudad: str) -> bool:
    from city_normalize import _clean
    return _clean(ciudad) in {_clean(a) for a in BOGOTA_ALIASES}


def build_pattern_masks(df: pd.DataFrame, patterns: list[dict]) -> dict[str, pd.Series]:
    """Construye máscaras booleanas con regex exacto por patrón."""
    masks = {}
    for p in patterns:
        label   = p["label"]
        pattern = p["pattern"]
        try:
            masks[label] = df["Guia#"].str.contains(pattern, regex=True, na=False)
        except re.error as e:
            logger.warning(f"Regex inválido para patrón '{label}': {e}")
            masks[label] = pd.Series(False, index=df.index)
    return masks


def any_pattern_mask(masks: dict[str, pd.Series]) -> pd.Series:
    if not masks:
        return pd.Series(False)
    result = pd.Series(False, index=next(iter(masks.values())).index)
    for m in masks.values():
        result |= m
    return result


# ── Formato de hoja ───────────────────────────────────────────────────────────

def style_sheet(ws, df: pd.DataFrame):
    """
    Aplica formato a la hoja.
    Las filas con ciudades ambiguas reciben fondo amarillo solo en la celda Ciudad.
    """
    if df.empty:
        ws.cell(row=1, column=1, value="Sin datos para esta sección")
        return

    cols = df.columns.tolist()
    city_col_idx = cols.index("Ciudad") + 1 if "Ciudad" in cols else None

    # Encabezados
    for ci, cn in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(cn, 18)

    # Datos
    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_vals   = list(row)
        base_fill  = ALT_FILL if ri % 2 == 0 else NORMAL_FILL

        # ¿Es ciudad ambigua?
        ciudad_val = ""
        if city_col_idx:
            ciudad_val = str(row_vals[city_col_idx - 1])
        row_is_ambig = is_ambiguous(ciudad_val)

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = BORDER
            c.alignment = LEFT

            if row_is_ambig and ci == city_col_idx:
                # Solo la celda de ciudad va en amarillo
                c.fill = AMBIG_FILL
                c.font = AMBIG_FONT
            else:
                c.fill = base_fill
                c.font = CELL_FONT

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_sheet(wb, name: str, df: pd.DataFrame):
    ws = wb.create_sheet(name[:31])
    style_sheet(ws, df)
    return ws


# ── Constructor principal ─────────────────────────────────────────────────────

def build_excel(df: pd.DataFrame, patterns: list[dict]) -> tuple[bytes, dict]:
    """
    Genera el Excel procesado con las pestañas:
      1. Original  (datos normalizados, ciudades ambiguas en amarillo)
      2. Bogotá    (sin patrones)
      3. Ciudades y Municipios (sin patrones)
      4..N. Patrón XXXXX (Bogotá + ciudades mezclados)

    Retorna (bytes_del_excel, stats_dict)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    mask_bog     = df["Ciudad"].apply(is_bogota)
    pat_masks    = build_pattern_masks(df, patterns)
    mask_any_pat = any_pattern_mask(pat_masks)

    df_bog_clean    = df[mask_bog & ~mask_any_pat]
    df_cities_clean = df[~mask_bog & ~mask_any_pat]

    add_sheet(wb, "Original",               df)
    add_sheet(wb, "Bogotá",                 df_bog_clean)
    add_sheet(wb, "Ciudades y Municipios",  df_cities_clean)

    pattern_stats = {}
    for p in patterns:
        label = p["label"]
        mask  = pat_masks.get(label, pd.Series(False, index=df.index))
        sub   = df[mask]
        if not sub.empty:
            add_sheet(wb, f"Patrón {label}", sub)
            pattern_stats[label] = len(sub)

    # Estadísticas de ciudades ambiguas detectadas
    ambig_cities = sorted({
        c for c in df["Ciudad"].unique() if is_ambiguous(c)
    })

    stats = {
        "total":           len(df),
        "bogota":          len(df_bog_clean),
        "cities":          len(df_cities_clean),
        "patterns":        pattern_stats,
        "ambig_cities":    ambig_cities,
        "ambig_count":     int(df["Ciudad"].apply(is_ambiguous).sum()),
    }

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), stats


# ── Lectura de archivo ────────────────────────────────────────────────────────

def read_excel_file(file_bytes: bytes) -> tuple[pd.DataFrame, dict[str, str]]:
    """
    Lee todas las hojas, usa la primera con columnas válidas.
    Normaliza ciudades automáticamente.
    """
    all_sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, dtype=str)
    target_df  = None
    target_map = None

    for sheet_name, df in all_sheets.items():
        df.columns = [str(c).strip() for c in df.columns]
        df.fillna("", inplace=True)
        try:
            col_map    = detect_columns(df)
            target_df  = normalize_df(df, col_map)
            target_map = col_map
            logger.info(f"Hoja usada: '{sheet_name}' — {len(df)} filas")
            break
        except ValueError:
            continue

    if target_df is None:
        raise ValueError(
            "No se encontró ninguna hoja con las columnas requeridas "
            f"(`{'`, `'.join(REQUIRED_COLUMNS)}`)."
        )

    return target_df, target_map


def make_summary(df: pd.DataFrame, patterns: list[dict], col_map: dict) -> str:
    mask_bog  = df["Ciudad"].apply(is_bogota)
    pat_masks = build_pattern_masks(df, patterns)
    mask_any  = any_pattern_mask(pat_masks)

    bog_n    = (mask_bog & ~mask_any).sum()
    cities_n = (~mask_bog & ~mask_any).sum()

    renames  = [f"`{v}` → `{k}`" for k, v in col_map.items() if k != v]
    col_info = ("\n🔄 *Columnas detectadas:* " + ", ".join(renames) + "\n") if renames else ""

    # Ciudades ambiguas detectadas
    ambig = [c for c in df["Ciudad"].unique() if is_ambiguous(c)]
    ambig_info = ""
    if ambig:
        ambig_info = f"\n⚠️ *Ciudades ambiguas (verificar departamento):* {', '.join(ambig)}\n"

    lines = [
        f"📦 *Total de despachos:* {len(df)}",
        col_info,
        f"🏙 *Bogotá (sin patrones):* {bog_n} guías",
        f"🗺 *Ciudades / Municipios (sin patrones):* {cities_n} guías",
        ambig_info,
        "📊 *Patrones detectados:*",
    ]
    for p in patterns:
        label = p["label"]
        cnt   = pat_masks.get(label, pd.Series(False, index=df.index)).sum()
        if cnt:
            lines.append(f"  • `{label}` → {cnt} guías")

    return "\n".join(l for l in lines if l is not None)
